import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import os
from openpyxl import load_workbook


def get_paths(folder, suffix):
    image_paths = []
    for file in folder.iterdir():
        if file.suffix == suffix:
            image_paths.append(file)
    return np.asarray(image_paths)


def data_loader(folder, split=None, patients=None):
    paths = sorted(Path(folder).glob("*.h5"))
    #paths = get_paths(folder, ".h5")
    measures_all = pd.DataFrame()
    for idx, path in enumerate(paths):
        if patients is not None and not int(path.name.split("_")[0]) in patients:
            continue
        print(f"Reading: {path.name} ({idx+1}/{len(paths)})")
        measures = pd.read_hdf(str(path), mode="r+", key="features")
        patient = measures.iloc[0]["sample"]
        if split is not None:
            if patient in split["train"]:
                measures["split"] = "train"
            elif patient in split["test"]:
                measures["split"] = "test"
            elif patient in split["val"]:
                measures["split"] = "val"
            else:
                raise RuntimeError(f"{patient} not present in train, val, test split.")
        measures_all = measures_all.append(measures)
    return measures_all


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, header=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet do not write the header since it already is there when the excel is created
    df.to_excel(writer, sheet_name, header, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def save_confusion_matrix(cm, diagnose_selection, classifier, save_path):
    save_path.parent.mkdir(parents=True, exist_ok=True)
    df_cm = pd.DataFrame(cm, columns=diagnose_selection, index=diagnose_selection)
    df_cm.index.name = 'Actual'
    df_cm.columns.name = 'Predicted'
    df_cm_sum = df_cm.sum(axis=1)
    df_cm_perc = df_cm.div(df_cm_sum, axis=0)*100
    annot = np.empty_like(df_cm).astype(str)
    for index in np.ndindex(annot.shape):
        annot[index[0], index[1]] = f"{df_cm_perc.iloc[index[0], index[1]]:.2f}%\n{df_cm.iloc[index[0], index[1]]:.0f}/{df_cm_sum.iloc[index[0]]:.0f}"

    hmap_plot = sns.heatmap(df_cm_perc, annot=annot, fmt='', linewidth=0.5, cmap="Blues", vmin=0, vmax=100)
    hmap_plot.set(title="Confusion matrix " + classifier)
    hmap_fig = hmap_plot.get_figure()
    hmap_fig.savefig(save_path)
    plt.close(hmap_fig)